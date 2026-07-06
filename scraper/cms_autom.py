import pandas as pd
import os
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Side, Font
from openpyxl.utils.dataframe import dataframe_to_rows

# # Define the directory to save the Excel files
# save_dir = '/path/to/save/directory'
# os.makedirs(save_dir, exist_ok=True)

# Define the path to the input Excel file
csv_path = 'CMS_REPORT.csv'

# Load the Excel file into a DataFrame
df = pd.read_csv(csv_path)

# Get unique ABN.TYPE values
abn_types = df['ABN.TYPE'].unique()

# Get the previous date
previous_date = (datetime.now() - timedelta(days=1)).strftime('%d-%m-%Y')
save_dir = os.path.join('Date_wise', previous_date)
os.makedirs(save_dir, exist_ok=True)

# Create thin border style
thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                     top=Side(style='thin'), bottom=Side(style='thin'))

# Define the Aptos Display font
aptos_font = Font(name='Aptos Display', size=11)

# Function to calculate column width based on content length
def calculate_column_width(value, rotated=True, min_width=1):
    # Estimate width based on content length
    estimated_width = max(len(str(value)), min_width)
    # Apply adjustment if text is rotated
    if rotated:
        # estimated_width = max(estimated_width, min_width)
        estimated_width = 1
    return estimated_width

# Save separate Excel files for each ABN.TYPE with formatting
for abn_type in abn_types:
    abn_df = df[df['ABN.TYPE'] == abn_type].copy()
    
    # Reset SNO. column starting from 1
    abn_df['SNO.'] = range(1, len(abn_df) + 1)
    
    # Create a new Workbook and select the active worksheet
    wb = Workbook()
    ws = wb.active
    
    # Append the dataframe rows to the worksheet
    for r_idx, row in enumerate(dataframe_to_rows(abn_df, index=False, header=True), 1):
        ws.append(row)
        for c_idx, cell in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx)
            cell.font = aptos_font  # Apply the Aptos Display font
            cell.alignment = Alignment(wrap_text=True, vertical='center', horizontal='center')
            cell.border = thin_border
            if r_idx == 1 or c_idx <= 24 or c_idx == len(row):  # Header or columns till 'DIV' or Reporting Date
                cell.alignment = Alignment(text_rotation=90, wrap_text=True, vertical='center', horizontal='center')
                # Adjust column width considering rotation
                ws.column_dimensions[cell.column_letter].width = calculate_column_width(cell.value, rotated=True)
            else:
                # Adjust column width without considering rotation
                ws.column_dimensions[cell.column_letter].width = calculate_column_width(cell.value, rotated=False)
    
    # Set minimum row height according to the content with a minimum of 90 pixels
    for row in ws.iter_rows(min_row=1, max_col=abn_df.shape[1], max_row=ws.max_row):
        for cell in row:
            if ws.row_dimensions[cell.row].height:
                ws.row_dimensions[cell.row].height = max(ws.row_dimensions[cell.row].height, 90)
            else:
                ws.row_dimensions[cell.row].height = 90
    
    
    # Adjust column widths to fit content
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter  # Get the column name
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length)  # Add a little extra space
        ws.column_dimensions[column].width = adjusted_width
    
    # Define the filename based on ABN.TYPE and the previous date
    excel_filename = f'ABNORMALITY {abn_type} {previous_date} 00.00 HRS TO 23.59 HRS.xlsx'
    excel_path = os.path.join(save_dir, excel_filename)
    
    # Save the workbook
    wb.save(excel_path)

print("Process completed successfully.")
